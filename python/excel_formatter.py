#!/usr/bin/env python3
"""
Excel Formatter Script - Apply conditional formatting and styling
Usage: python3 excel_formatter.py --file <path> --format <type> --range <range>
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import argparse
import sys
import json

def apply_conditional_formatting(file_path, format_type, range, sheet_name=None):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if format_type == "highlight-values-above-100":
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000')
            ws.conditional_formatting.add(range, CellIsRule(operator='greaterThan', formula=['100'], fill=red_fill))
            
        elif format_type == "highlight-values-below-50":
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00')
            ws.conditional_formatting.add(range, CellIsRule(operator='lessThan', formula=['50'], fill=yellow_fill))
            
        elif format_type == "color-scale":
            ws.conditional_formatting.add(range, ColorScaleRule(
                start_type='min', start_color='FF0000',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='00FF00'
            ))
            
        elif format_type == "data-bars":
            ws.conditional_formatting.add(range, DataBarRule(
                start_type='min', end_type='max', color='638EC6'
            ))
        
        wb.save(file_path)
        
        result = {
            "success": True,
            "message": f"Applied {format_type} formatting to range {range}",
            "file": file_path,
            "sheet": sheet_name or wb.active.title
        }
        
        return json.dumps(result)
    
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })

def apply_styling(file_path, style_type, range, sheet_name=None):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if style_type == "bold-headers":
            # Make first row bold
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
                
        elif style_type == "add-borders":
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws[range]:
                for cell in row:
                    cell.border = thin_border
                    
        elif style_type == "center-align":
            center_alignment = Alignment(horizontal='center', vertical='center')
            for row in ws[range]:
                for cell in row:
                    cell.alignment = center_alignment
        
        wb.save(file_path)
        
        result = {
            "success": True,
            "message": f"Applied {style_type} styling to range {range}",
            "file": file_path,
            "sheet": sheet_name or wb.active.title
        }
        
        return json.dumps(result)
    
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })

def create_chart(file_path, chart_type, range, sheet_name=None, position=None):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        # Parse range for chart data
        data_ref = Reference(ws, range_string=range)
        
        if chart_type == "bar":
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            raise ValueError(f"Unsupported chart type: {chart_type}")
        
        chart.add_data(data_ref, titles_from_data=True)
        
        # Position chart (default to E1)
        if position:
            ws.add_chart(chart, position)
        else:
            ws.add_chart(chart, "E1")
        
        wb.save(file_path)
        
        result = {
            "success": True,
            "message": f"Created {chart_type} chart from range {range}",
            "file": file_path,
            "sheet": sheet_name or wb.active.title,
            "chart_position": position or "E1"
        }
        
        return json.dumps(result)
    
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })

def main():
    parser = argparse.ArgumentParser(description='Apply Excel formatting and styling')
    parser.add_argument('--file', required=True, help='Excel file path')
    parser.add_argument('--action', choices=['conditional-format', 'style', 'chart'], required=True, help='Action to perform')
    parser.add_argument('--type', required=True, help='Format/style/chart type')
    parser.add_argument('--range', required=True, help='Cell range (e.g., A1:C10)')
    parser.add_argument('--sheet', help='Sheet name')
    parser.add_argument('--position', help='Chart position (e.g., E1)')
    
    args = parser.parse_args()
    
    if args.action == 'conditional-format':
        result = apply_conditional_formatting(args.file, args.type, args.range, args.sheet)
    elif args.action == 'style':
        result = apply_styling(args.file, args.type, args.range, args.sheet)
    elif args.action == 'chart':
        result = create_chart(args.file, args.type, args.range, args.sheet, args.position)
    
    print(result)

if __name__ == "__main__":
    main()
