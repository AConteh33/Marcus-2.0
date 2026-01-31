#!/usr/bin/env python3
"""
Excel Writer Script - Create and update Excel files
Usage: python3 excel_writer.py --file <path> --data <json_data> --action <create|update>
"""

import pandas as pd
import openpyxl
import argparse
import sys
import json
from pathlib import Path

def create_excel_file(file_path, data, sheet_name="Sheet1"):
    try:
        # Convert data to DataFrame
        if isinstance(data, list):
            df = pd.DataFrame(data)
        elif isinstance(data, dict):
            df = pd.DataFrame([data])
        else:
            df = pd.DataFrame(data)
        
        # Create Excel file
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        result = {
            "success": True,
            "message": f"Excel file created: {file_path}",
            "rows": len(df),
            "columns": len(df.columns),
            "sheet": sheet_name
        }
        
        return json.dumps(result)
    
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })

def update_excel_file(file_path, data, sheet_name=None, cell=None):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        
        if cell and isinstance(data, (str, int, float)):
            # Update single cell
            ws[cell] = data
            message = f"Updated cell {cell} with value: {data}"
        else:
            # Update range or add new data
            if isinstance(data, list):
                for row_idx, row_data in enumerate(data, start=1):
                    for col_idx, value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                message = f"Updated data range with {len(data)} rows"
            else:
                message = f"Updated sheet with data"
        
        wb.save(file_path)
        
        result = {
            "success": True,
            "message": message,
            "file": file_path,
            "sheet": sheet_name or wb.active.title
        }
        
        return json.dumps(result)
    
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": str(e)
        })

def main():
    parser = argparse.ArgumentParser(description='Create or update Excel files')
    parser.add_argument('--file', required=True, help='Excel file path')
    parser.add_argument('--data', required=True, help='JSON data string')
    parser.add_argument('--action', choices=['create', 'update'], required=True, help='Action to perform')
    parser.add_argument('--sheet', help='Sheet name')
    parser.add_argument('--cell', help='Specific cell to update (e.g., A1)')
    
    args = parser.parse_args()
    
    # Parse JSON data
    try:
        data = json.loads(args.data)
    except json.JSONDecodeError:
        result = {"success": False, "error": "Invalid JSON data"}
        print(json.dumps(result))
        return
    
    if args.action == 'create':
        result = create_excel_file(args.file, data, args.sheet)
    else:
        result = update_excel_file(args.file, data, args.sheet, args.cell)
    
    print(result)

if __name__ == "__main__":
    main()
